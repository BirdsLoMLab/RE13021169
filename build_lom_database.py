#!/usr/bin/env python3
"""
build_lom_database.py — Turn decoded config tables into a queryable SQLite database
and a human-readable items/mounts workbook.

Input:  a directory of decoded tables produced by decode_config_data.py
        (one <Table>.json per table, plus _index.json).
Output:
  <out>/lom_config.sqlite   — every table as a SQL table (named columns, arrays
                              stored as JSON text), plus:
                                _tables      (name, records, has_schema)
                                _fields      (table, idx, field, type, xor)
                                names_en     (id, text, source)  — Language_en + Language_ui_en
                              and a *_named view for every table whose `name`
                              column resolves in names_en.
  <out>/LOM_Items_Mounts_Database.xlsx — curated sheets with English names.

Usage:
  python3 build_lom_database.py data/tables --out dist
  python3 build_lom_database.py data/tables --out dist --schema-dir data/schemas
"""

import argparse
import json
import os
import sqlite3
import sys

CURATED_SHEETS = [
    # (sheet title, table, name field)
    ("Items (Goods)", "Goods", "name"),
    ("Mounts", "Mount", "name"),
    ("Mount Skins", "Mount_skin", None),
    ("Mount Levels", "Mount_level", None),
    ("Mount Abilities", "Mount_ability", None),
    ("Equipment", "Equipment", "name"),
    ("Equipment Suits", "Equipment_suit", None),
    ("Artifacts", "Artifact", "name"),
    ("Artifact Skins", "Artifact_skin", None),
    ("Back Decorations", "Back_decoration", "name"),
    ("Back Skins", "Back_skin", None),
    ("Pets (Pals)", "Pet", "name"),
    ("Avians (Fly)", "Fly", "name"),
    ("Angels (Star Heroes)", "Angel", "name"),
    ("Spirits", "Spirit", "mame"),
    ("Relics", "Relic", "name"),
    ("Jobs (Classes)", "Jobs", "name"),
    ("Attributes", "Attribute", "name"),
    ("Skills", "Skill", "name"),
    ("Buffs", "Buff", "name"),
    ("Fate Cards", "Fate", "name"),
    ("Rings", "Ring", "name"),
    ("Badges", "Badge", "name"),
    ("Levels (PvP factor)", "Level", None),
]


def load_table(tables_dir, name):
    path = os.path.join(tables_dir, f"{name}.json")
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_names(tables_dir):
    names = {}
    for src in ("Language_en", "Language_ui_en"):
        rows = load_table(tables_dir, src) or []
        for r in rows:
            if isinstance(r, dict):
                rid, text = r.get("id"), r.get("text")
            else:
                rid, text = (r[0], r[1]) if len(r) >= 2 else (None, None)
            if rid is not None and rid not in names:
                names[rid] = (text, src)
    return names


def sql_ident(name):
    return '"' + str(name).replace('"', '""') + '"'


def cell(v):
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False, separators=(",", ":"))
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int) and abs(v) > 9223372036854775807:
        return str(v)  # beyond SQLite/Excel int range (decoded BigNumber)
    return v


def build_sqlite(tables_dir, schema_dir, out_path, names):
    if os.path.exists(out_path):
        os.remove(out_path)
    con = sqlite3.connect(out_path)
    cur = con.cursor()
    cur.execute("PRAGMA journal_mode=OFF")
    cur.execute("PRAGMA synchronous=OFF")

    cur.execute("CREATE TABLE _tables (name TEXT PRIMARY KEY, records INTEGER, has_schema INTEGER, columns TEXT)")
    cur.execute("CREATE TABLE _fields (tbl TEXT, idx INTEGER, field TEXT, type TEXT, xor INTEGER)")
    cur.execute("CREATE TABLE names_en (id INTEGER PRIMARY KEY, text TEXT, source TEXT)")
    cur.executemany("INSERT INTO names_en VALUES (?,?,?)", [(k, v[0], v[1]) for k, v in names.items()])

    index = load_table(tables_dir, "_index") or {}
    views = 0
    for fname in sorted(os.listdir(tables_dir)):
        if not fname.endswith(".json") or fname.startswith("_"):
            continue
        tname = fname[:-5]
        rows = load_table(tables_dir, tname)
        if rows is None:
            continue
        # Column set = union of keys in insertion order; raw list rows get c0..cN
        cols = []
        seen = set()
        for r in rows:
            if isinstance(r, dict):
                keys = r.keys()
            else:
                keys = [f"c{i}" for i in range(len(r))]
            for k in keys:
                if k not in seen:
                    seen.add(k)
                    cols.append(k)
        if not cols:
            cols = ["c0"]
        cur.execute(f"CREATE TABLE {sql_ident(tname)} ({', '.join(sql_ident(c) for c in cols)})")
        data = []
        for r in rows:
            if isinstance(r, dict):
                data.append([cell(r.get(c)) for c in cols])
            else:
                data.append([cell(r[i]) if i < len(r) else None for i in range(len(cols))])
        cur.executemany(
            f"INSERT INTO {sql_ident(tname)} VALUES ({','.join('?' * len(cols))})", data
        )
        meta = index.get(tname, {})
        cur.execute(
            "INSERT INTO _tables VALUES (?,?,?,?)",
            (tname, len(rows), int(bool(meta.get("has_schema", isinstance(rows[0], dict) if rows else 0))), json.dumps(cols)),
        )
        # Schema fields
        schema_path = os.path.join(schema_dir, f"Config{tname}.json") if schema_dir else None
        if schema_path and os.path.exists(schema_path):
            with open(schema_path) as f:
                sch = json.load(f)
            cur.executemany(
                "INSERT INTO _fields VALUES (?,?,?,?,?)",
                [(tname, fl.get("index"), fl.get("name"), fl.get("type"), int(bool(fl.get("xor")))) for fl in sch.get("fields", [])],
            )
        # Named view
        name_col = "name" if "name" in cols else ("mame" if "mame" in cols else None)
        if name_col:
            cur.execute(
                f"CREATE VIEW {sql_ident(tname + '_named')} AS "
                f"SELECT n.text AS name_en, t.* FROM {sql_ident(tname)} t "
                f"LEFT JOIN names_en n ON n.id = t.{sql_ident(name_col)}"
            )
            views += 1
    con.commit()
    con.close()
    return views


def build_xlsx(tables_dir, out_path, names):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping xlsx (pip install openpyxl)")
        return False

    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E78")

    # README sheet
    ws = wb.create_sheet("README")
    readme = [
        ["LOM Config Database — curated export"],
        ["Source", "bundle-firstload-res config/datas (FilePack v5, 908 tables), capture 2026-02-28"],
        ["Decoder", "decode_config_data.py (byte XOR 255&~(32^b) → zlib → records; protected ints ^ 24455)"],
        ["Names", "name_en resolved from Language_en, then Language_ui_en, by the `name` id"],
        ["Arrays", "Nested arrays/objects are stored as JSON text in a single cell"],
        ["Attr pairs", "[[attrId, value], ...] — attrId per Attribute sheet; rates are /10000 unless num_type says otherwise"],
        ["Credit", "Bird → Discord @birrrd08"],
        [],
        ["Sheet", "Table", "Rows"],
    ]
    for r in readme:
        ws.append(r)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100

    for title, tname, name_field in CURATED_SHEETS:
        rows = load_table(tables_dir, tname)
        if not rows:
            ws.append([title, tname, "MISSING"])
            continue
        sheet = wb.create_sheet(title[:31])
        cols = []
        seen = set()
        for r in rows:
            for k in (r.keys() if isinstance(r, dict) else range(len(r))):
                if k not in seen:
                    seen.add(k)
                    cols.append(k)
        header = (["name_en"] if name_field else []) + [str(c) for c in cols]
        sheet.append(header)
        for c in sheet[1]:
            c.font = head_font
            c.fill = head_fill
        for r in rows:
            out = []
            if name_field:
                nid = r.get(name_field) if isinstance(r, dict) else None
                out.append(names.get(nid, (None,))[0] if nid is not None else None)
            for c in cols:
                v = r.get(c) if isinstance(r, dict) else (r[c] if c < len(r) else None)
                v = cell(v)
                if isinstance(v, str) and len(v) > 32000:
                    v = v[:32000] + "…"
                out.append(v)
            sheet.append(out)
        sheet.freeze_panes = "B2" if name_field else "A2"
        for i, h in enumerate(header, 1):
            sheet.column_dimensions[get_column_letter(i)].width = min(max(12, len(h) + 2), 40)
        ws.append([title, tname, len(rows)])

    wb.save(out_path)
    return True


def main():
    ap = argparse.ArgumentParser(description="Build SQLite + XLSX from decoded LOM config tables")
    ap.add_argument("tables_dir", help="Directory with decoded <Table>.json files")
    ap.add_argument("--out", default="dist", help="Output directory (default: dist)")
    ap.add_argument("--schema-dir", default="data/schemas", help="Config*.json schema directory")
    ap.add_argument("--no-xlsx", action="store_true")
    ap.add_argument("--no-sqlite", action="store_true")
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)
    names = load_names(args.tables_dir)
    print(f"Loaded {len(names):,} English strings")

    if not args.no_sqlite:
        sq = os.path.join(args.out, "lom_config.sqlite")
        views = build_sqlite(args.tables_dir, args.schema_dir, sq, names)
        print(f"SQLite: {sq} ({os.path.getsize(sq)/1e6:.1f} MB, {views} *_named views)")
    if not args.no_xlsx:
        xl = os.path.join(args.out, "LOM_Items_Mounts_Database.xlsx")
        if build_xlsx(args.tables_dir, xl, names):
            print(f"XLSX: {xl} ({os.path.getsize(xl)/1e6:.1f} MB)")


if __name__ == "__main__":
    sys.exit(main())
